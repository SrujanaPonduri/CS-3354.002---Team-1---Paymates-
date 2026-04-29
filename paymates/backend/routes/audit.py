# routes/audit.py
# Responsible for: UC-12 — Audit Expenses and Budget.
# Aagam Shah
#
# FR-24: Visual representations of expense trends over time.
# FR-25: Visual comparison of current expenditures against allocated budgets.
# FR-26: Key financial metrics and summary reports (total spending, top categories).
# FR-27: Export expense audit reports in CSV / XLSX format.
# NFR-08: All search/report operations must complete within 2 seconds.
#
# Endpoints:
#   GET  /api/homes/<home_id>/audit/summary    — FR-26 aggregate metrics
#   GET  /api/homes/<home_id>/audit/trends     — FR-24 daily spending time-series
#   GET  /api/homes/<home_id>/audit/budget-vs-actual  — FR-25 budget vs spend
#   GET  /api/homes/<home_id>/audit/export/csv — FR-27 CSV download
#   GET  /api/homes/<home_id>/audit/export/xlsx— FR-27 XLSX download

import csv
import io
from collections import defaultdict
from datetime import date, timedelta
from flask import Blueprint, jsonify, request, Response
from mock_db import DB

audit_bp = Blueprint("audit", __name__)

# ── Budgets in-memory store (FR-04 / FR-05 / FR-06) ─────────────────────────
# We keep budgets inside mock_db's DB dict under key "budgets".
# Schema:
#   {
#     id            : str
#     home_id       : str
#     category      : str
#     budget_amount : float
#     month         : int  (1-12)
#     year          : int
#     visibility    : str  "private" | "group" | "all"
#     created_by    : str  user_id
#   }
if "budgets" not in DB:
    DB["budgets"] = {}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_period(args) -> tuple[str, str]:
    """Return (start_date, end_date) from query params or sensible defaults.

    Supported ?period= values:
      this_month, last_month, last_3_months, last_6_months, ytd, custom
    For custom, caller must also pass ?start_date= and ?end_date=.
    """
    period     = (args.get("period") or "this_month").strip().lower()
    today      = date.today()
    first_this = today.replace(day=1)

    if period == "this_month":
        start = first_this.isoformat()
        # Use last day of this month so bills/expenses dated today or later
        # in the same calendar month are included (not cut off at today).
        import calendar
        last_day = calendar.monthrange(today.year, today.month)[1]
        end   = today.replace(day=last_day).isoformat()
    elif period == "last_month":
        last_month_end   = first_this - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        start = last_month_start.isoformat()
        end   = last_month_end.isoformat()
    elif period == "last_3_months":
        start = (today - timedelta(days=90)).isoformat()
        end   = today.isoformat()
    elif period == "last_6_months":
        start = (today - timedelta(days=180)).isoformat()
        end   = today.isoformat()
    elif period == "ytd":
        start = today.replace(month=1, day=1).isoformat()
        end   = today.isoformat()
    else:  # custom
        start = (args.get("start_date") or first_this.isoformat()).strip()
        end   = (args.get("end_date")   or today.isoformat()).strip()

    return start, end


def _bills_in_period(home_id: str, start: str, end: str) -> list[dict]:
    return [
        b for b in DB["bills"].values()
        if b["home_id"] == home_id
        and start <= b.get("date", "9999") <= end
    ]


def _expenses_in_period(home_id: str, start: str, end: str) -> list[dict]:
    return [
        e for e in DB["expenses"].values()
        if e["home_id"] == home_id
        and start <= e.get("start_date", "9999") <= end
    ]


def _items_in_period(home_id: str, start: str, end: str) -> list[dict]:
    """Return inventory items purchased within the period for this home."""
    return [
        i for i in DB["items"].values()
        if i.get("home_id") == home_id
        and start <= (i.get("purchased_on") or "9999") <= end
    ]


def _item_total(item: dict) -> float:
    """Line-total cost of an inventory item (quantity × unit_price)."""
    try:
        return round(float(item.get("quantity", 0) or 0)
                     * float(item.get("unit_price", 0) or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _category_totals(bills: list, expenses: list,
                     items: list | None = None) -> dict[str, float]:
    """Sum spend by category across bills, expenses and inventory items."""
    totals: dict[str, float] = defaultdict(float)
    for b in bills:
        totals[b.get("category") or "Uncategorised"] += b.get("total", 0)
    for e in expenses:
        totals[e.get("category") or e.get("title") or "Uncategorised"] += e.get("amount", 0)
    for i in (items or []):
        totals[i.get("category") or "Uncategorised"] += _item_total(i)
    return dict(totals)


# ---------------------------------------------------------------------------
# FR-26 — Aggregate summary metrics
# ---------------------------------------------------------------------------
# GET /api/homes/<home_id>/audit/summary
@audit_bp.route("/homes/<home_id>/audit/summary", methods=["GET"])
def audit_summary(home_id):
    """FR-26: Return key financial metrics for the requested period.

    Returns:
      total_spending      — sum of bill totals + expense amounts + item costs
      total_bills         — count of bills
      total_expenses      — count of expenses
      total_items         — count of inventory items purchased in period
      total_transactions  — bills + expenses + items
      recurring_expenses  — count of recurring expenses
      top_category        — category with highest spend
      category_breakdown  — dict of category → total spend
      period              — { start_date, end_date }
    """
    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    start, end  = _parse_period(request.args)
    bills       = _bills_in_period(home_id, start, end)
    expenses    = _expenses_in_period(home_id, start, end)
    items       = _items_in_period(home_id, start, end)

    cat_totals  = _category_totals(bills, expenses, items)
    top_cat     = max(cat_totals, key=cat_totals.get) if cat_totals else None
    total_spend = round(sum(cat_totals.values()), 2)

    # Sort breakdown by value descending for the frontend chart
    sorted_breakdown = dict(
        sorted(cat_totals.items(), key=lambda kv: kv[1], reverse=True)
    )

    recurring_count = sum(
        1 for e in expenses if e.get("expense_type") == "recurring"
    )

    return jsonify({
        "total_spending":     total_spend,
        "total_bills":        len(bills),
        "total_expenses":     len(expenses),
        "total_items":        len(items),
        "total_transactions": len(bills) + len(expenses) + len(items),
        "recurring_expenses": recurring_count,
        "top_category":       top_cat,
        "top_category_amount": round(cat_totals.get(top_cat, 0), 2) if top_cat else 0,
        "top_category_pct":   round(cat_totals.get(top_cat, 0) / total_spend * 100, 1)
                               if total_spend else 0,
        "category_breakdown": sorted_breakdown,
        "period": {"start_date": start, "end_date": end},
    }), 200


# ---------------------------------------------------------------------------
# FR-24 — Daily spending trend (time-series)
# ---------------------------------------------------------------------------
# GET /api/homes/<home_id>/audit/trends
@audit_bp.route("/homes/<home_id>/audit/trends", methods=["GET"])
def audit_trends(home_id):
    """FR-24: Return daily spend aggregated into a time-series array.

    Each element:  { date: "YYYY-MM-DD", amount: float, is_today: bool }

    Suitable for bar/line charts in the frontend.
    """
    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    start, end  = _parse_period(request.args)
    bills       = _bills_in_period(home_id, start, end)
    expenses    = _expenses_in_period(home_id, start, end)
    items       = _items_in_period(home_id, start, end)

    # Aggregate by date
    daily: dict[str, float] = defaultdict(float)
    for b in bills:
        d = b.get("date", "")
        if d:
            daily[d] += b.get("total", 0)
    for e in expenses:
        d = e.get("start_date", "")
        if d:
            daily[d] += e.get("amount", 0)
    for i in items:
        d = i.get("purchased_on", "")
        if d:
            daily[d] += _item_total(i)

    # Build a full day-by-day series (zero-fill gaps)
    today_str   = date.today().isoformat()
    try:
        cursor  = date.fromisoformat(start)
        end_d   = date.fromisoformat(end)
    except ValueError:
        return jsonify({"error": "Invalid date range"}), 400

    series = []
    while cursor <= end_d:
        ds = cursor.isoformat()
        series.append({
            "date":     ds,
            "amount":   round(daily.get(ds, 0), 2),
            "is_today": ds == today_str,
        })
        cursor += timedelta(days=1)

    return jsonify({
        "trends": series,
        "period": {"start_date": start, "end_date": end},
    }), 200


# ---------------------------------------------------------------------------
# FR-25 — Budget vs. Actual comparison
# ---------------------------------------------------------------------------
# GET /api/homes/<home_id>/audit/budget-vs-actual
@audit_bp.route("/homes/<home_id>/audit/budget-vs-actual", methods=["GET"])
def budget_vs_actual(home_id):
    """FR-25: Compare each budget category against actual spend for the period.

    Each row:
      category        — category name
      budget_amount   — allocated amount (0 if no budget defined)
      actual_spent    — actual spend in the period
      remaining       — budget_amount - actual_spent  (can be negative)
      over_budget     — bool
      pct_used        — percentage of budget consumed (null if no budget)
    """
    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    start, end = _parse_period(request.args)
    bills      = _bills_in_period(home_id, start, end)
    expenses   = _expenses_in_period(home_id, start, end)
    items      = _items_in_period(home_id, start, end)

    # Period month/year for matching budgets
    try:
        period_start = date.fromisoformat(start)
        month_f      = period_start.month
        year_f       = period_start.year
    except ValueError:
        month_f = year_f = None

    # Actual spend per category — bills + expenses + inventory items
    actual: dict[str, float] = defaultdict(float)
    for b in bills:
        actual[b.get("category") or "Uncategorised"] += b.get("total", 0)
    for e in expenses:
        actual[e.get("category") or e.get("title") or "Uncategorised"] += e.get("amount", 0)
    for i in items:
        actual[i.get("category") or "Uncategorised"] += _item_total(i)

    # Budgets for this home (filter by month/year only when budget specifies them).
    # Supports both schemas in DB["budgets"]:
    #   • UC-12 audit:   budget_amount + month + year
    #   • UC-11 budgets: limit (no month/year)
    budgets_for_home: dict[str, float] = {}
    for bgt in DB["budgets"].values():
        if bgt["home_id"] != home_id:
            continue
        if month_f and bgt.get("month") and bgt.get("month") != month_f:
            continue
        if year_f and bgt.get("year") and bgt.get("year") != year_f:
            continue
        amt = bgt.get("budget_amount", bgt.get("limit", 0)) or 0
        budgets_for_home[bgt["category"]] = amt

    # Merge all categories
    all_cats = set(actual.keys()) | set(budgets_for_home.keys())
    rows = []
    over_budget_count = 0
    for cat in sorted(all_cats):
        budget_amt = budgets_for_home.get(cat, 0)
        spent      = round(actual.get(cat, 0), 2)
        remaining  = round(budget_amt - spent, 2)
        over       = budget_amt > 0 and spent > budget_amt
        if over:
            over_budget_count += 1
        pct = round(spent / budget_amt * 100, 1) if budget_amt > 0 else None
        rows.append({
            "category":     cat,
            "budget_amount": round(budget_amt, 2),
            "actual_spent": spent,
            "remaining":    remaining,
            "over_budget":  over,
            "pct_used":     pct,
        })

    total_budget = round(sum(budgets_for_home.values()), 2)
    total_spent  = round(sum(actual.values()), 2)
    # Spend that falls inside categories the user has actually budgeted —
    # use this for the audit "Budget Used" metric so spending in unbudgeted
    # categories does not skew the percentage.
    budgeted_actual_spent = round(
        sum(actual.get(cat, 0) for cat in budgets_for_home), 2
    )
    budget_used_pct = round(budgeted_actual_spent / total_budget * 100, 1) \
        if total_budget > 0 else None

    return jsonify({
        "rows":                  rows,
        "total_budget":          total_budget,
        "total_spent":           total_spent,
        "budgeted_actual_spent": budgeted_actual_spent,
        "budget_used_pct":       budget_used_pct,
        "over_budget_count":     over_budget_count,
        "period": {"start_date": start, "end_date": end},
    }), 200


# ---------------------------------------------------------------------------
# Unified transaction history — bills + expenses + items in one feed
# ---------------------------------------------------------------------------
# GET /api/homes/<home_id>/audit/transactions
@audit_bp.route("/homes/<home_id>/audit/transactions", methods=["GET"])
def audit_transactions(home_id):
    """Return a unified transaction list (bills + expenses + inventory items).

    Query params:
      period, start_date, end_date — same as other audit endpoints
      per_page                     — max records returned (default 50, max 200)

    Each record shape:
      { id, type, title, category, date, amount, by_id, by_name, extra }
      type ∈ "bill" | "expense" | "item"
    """
    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    try:
        per_page = min(200, max(1, int(request.args.get("per_page", 50))))
    except (TypeError, ValueError):
        per_page = 50

    start, end = _parse_period(request.args)
    bills      = _bills_in_period(home_id, start, end)
    expenses   = _expenses_in_period(home_id, start, end)
    items      = _items_in_period(home_id, start, end)

    transactions: list[dict] = []

    for b in bills:
        creator = DB["users"].get(b.get("created_by") or "", {})
        transactions.append({
            "id":       b.get("id"),
            "type":     "bill",
            "title":    b.get("title", "—"),
            "category": b.get("category") or "Uncategorised",
            "date":     b.get("date") or "",
            "amount":   round(float(b.get("total", 0) or 0), 2),
            "by_id":    b.get("created_by"),
            "by_name":  creator.get("name", "—"),
            "extra":    b.get("split_type") or "",
        })

    for e in expenses:
        creator = DB["users"].get(e.get("creator_id") or "", {})
        transactions.append({
            "id":       e.get("id"),
            "type":     "expense",
            "title":    e.get("title", "—"),
            "category": e.get("category") or e.get("expense_type") or "Expense",
            "date":     e.get("start_date") or "",
            "amount":   round(float(e.get("amount", 0) or 0), 2),
            "by_id":    e.get("creator_id"),
            "by_name":  creator.get("name", "—"),
            "extra":    e.get("frequency") or "one-time",
        })

    for i in items:
        owner_id = (i.get("owners") or [None])[0]
        owner    = DB["users"].get(owner_id or "", {})
        qty      = float(i.get("quantity", 0) or 0)
        price    = float(i.get("unit_price", 0) or 0)
        transactions.append({
            "id":       i.get("id"),
            "type":     "item",
            "title":    i.get("name", "—"),
            "category": i.get("category") or "Uncategorised",
            "date":     i.get("purchased_on") or "",
            "amount":   _item_total(i),
            "by_id":    owner_id,
            "by_name":  owner.get("name", "—"),
            "extra":    f"qty {qty:g} × ${price:.2f}",
        })

    transactions.sort(key=lambda r: r["date"], reverse=True)
    total = len(transactions)

    return jsonify({
        "transactions": transactions[:per_page],
        "total":        total,
        "period": {"start_date": start, "end_date": end},
    }), 200


# ---------------------------------------------------------------------------
# FR-27 — Export CSV
# ---------------------------------------------------------------------------
# GET /api/homes/<home_id>/audit/export/csv
@audit_bp.route("/homes/<home_id>/audit/export/csv", methods=["GET"])
def export_csv(home_id):
    """FR-27: Export all bills and expenses in the period as a CSV file."""
    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    start, end = _parse_period(request.args)
    bills      = _bills_in_period(home_id, start, end)
    expenses   = _expenses_in_period(home_id, start, end)
    items      = _items_in_period(home_id, start, end)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Type", "Title", "Category", "Amount", "Split Type / Frequency"])

    for b in sorted(bills, key=lambda x: x.get("date", ""), reverse=True):
        writer.writerow([
            b.get("date", ""),
            "Bill",
            b.get("title", ""),
            b.get("category", ""),
            f"{b.get('total', 0):.2f}",
            b.get("split_type", ""),
        ])

    for e in sorted(expenses, key=lambda x: x.get("start_date", ""), reverse=True):
        writer.writerow([
            e.get("start_date", ""),
            "Expense",
            e.get("title", ""),
            e.get("category") or e.get("expense_type", ""),
            f"{e.get('amount', 0):.2f}",
            e.get("frequency") or "one-time",
        ])

    for i in sorted(items, key=lambda x: x.get("purchased_on", ""), reverse=True):
        qty   = float(i.get("quantity", 0) or 0)
        price = float(i.get("unit_price", 0) or 0)
        writer.writerow([
            i.get("purchased_on", ""),
            "Item",
            i.get("name", ""),
            i.get("category", ""),
            f"{_item_total(i):.2f}",
            f"qty {qty:g} × ${price:.2f}",
        ])

    csv_bytes = output.getvalue().encode("utf-8")
    filename  = f"paymates_audit_{start}_to_{end}.csv"
    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# FR-27 — Export XLSX
# ---------------------------------------------------------------------------
# GET /api/homes/<home_id>/audit/export/xlsx
@audit_bp.route("/homes/<home_id>/audit/export/xlsx", methods=["GET"])
def export_xlsx(home_id):
    """FR-27: Export all bills and expenses in the period as an XLSX file.

    Requires openpyxl (listed in requirements.txt).  Falls back to a 501
    error if the package is not available.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server"}), 501

    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    home_name  = DB["homes"][home_id].get("name", "Home")
    start, end = _parse_period(request.args)
    bills      = _bills_in_period(home_id, start, end)
    expenses   = _expenses_in_period(home_id, start, end)
    items      = _items_in_period(home_id, start, end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Audit"

    # ── Header row ────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="7C3AED")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    headers = ["Date", "Type", "Title", "Category", "Amount ($)", "Split / Frequency"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ─────────────────────────────────────────────────────────
    row_num = 2
    BILL_FILL    = PatternFill("solid", fgColor="1C1C2E")   # dark purple tint
    EXPENSE_FILL = PatternFill("solid", fgColor="0F1530")

    for b in sorted(bills, key=lambda x: x.get("date", ""), reverse=True):
        ws.append([
            b.get("date", ""),
            "Bill",
            b.get("title", ""),
            b.get("category", ""),
            round(b.get("total", 0), 2),
            b.get("split_type", ""),
        ])
        row_num += 1

    for e in sorted(expenses, key=lambda x: x.get("start_date", ""), reverse=True):
        ws.append([
            e.get("start_date", ""),
            "Expense",
            e.get("title", ""),
            e.get("category") or e.get("expense_type", ""),
            round(e.get("amount", 0), 2),
            e.get("frequency") or "one-time",
        ])
        row_num += 1

    for i in sorted(items, key=lambda x: x.get("purchased_on", ""), reverse=True):
        qty   = float(i.get("quantity", 0) or 0)
        price = float(i.get("unit_price", 0) or 0)
        ws.append([
            i.get("purchased_on", ""),
            "Item",
            i.get("name", ""),
            i.get("category", ""),
            _item_total(i),
            f"qty {qty:g} × ${price:.2f}",
        ])
        row_num += 1

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Home", home_name])
    ws2.append(["Period", f"{start} to {end}"])
    ws2.append(["Total Bills", len(bills)])
    ws2.append(["Total Expenses", len(expenses)])
    ws2.append(["Total Items", len(items)])
    ws2.append(["Total Transactions", len(bills) + len(expenses) + len(items)])
    ws2.append(["Total Spend ($)", round(
        sum(b.get("total", 0) for b in bills) +
        sum(e.get("amount", 0) for e in expenses) +
        sum(_item_total(i) for i in items), 2
    )])

    # Auto-width columns (main sheet)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    # ── Stream to response ────────────────────────────────────────────────
    buf      = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"paymates_audit_{start}_to_{end}.xlsx"
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# UC-12 helper — Budget CRUD (FR-04 / FR-05 / FR-06)
# ---------------------------------------------------------------------------
# POST /api/homes/<home_id>/budgets
@audit_bp.route("/homes/<home_id>/budgets", methods=["POST"])
def create_budget(home_id):
    """FR-04/05/06: Create a budget for a home category.

    Body: { creator_id, category, budget_amount, month, year, visibility }
    """
    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    from mock_db import new_id
    data = request.get_json(silent=True) or {}

    creator_id    = (data.get("creator_id") or "").strip()
    category      = (data.get("category") or "").strip()
    visibility    = (data.get("visibility") or "all").strip().lower()
    today         = date.today()

    try:
        budget_amount = float(data.get("budget_amount", 0))
    except (TypeError, ValueError):
        return jsonify({"error": "budget_amount must be a number"}), 400

    if not category:
        return jsonify({"error": "category is required"}), 400
    if budget_amount <= 0:
        return jsonify({"error": "budget_amount must be greater than 0"}), 400
    if visibility not in ("private", "group", "all"):
        return jsonify({"error": "visibility must be private, group, or all"}), 400

    try:
        month = int(data.get("month", today.month))
        year  = int(data.get("year",  today.year))
    except (TypeError, ValueError):
        return jsonify({"error": "month and year must be integers"}), 400

    bgt_id = new_id()
    budget = {
        "id":              bgt_id,
        "home_id":         home_id,
        "category":        category,
        "budget_amount":   budget_amount,
        "current_balance": 0.0,
        "month":           month,
        "year":            year,
        "visibility":      visibility,
        "created_by":      creator_id,
    }
    DB["budgets"][bgt_id] = budget
    return jsonify({"budget": budget}), 201


# GET /api/homes/<home_id>/budgets
@audit_bp.route("/homes/<home_id>/budgets", methods=["GET"])
def list_budgets(home_id):
    """FR-04/05: List all budgets for a home with live-computed current_balance.

    Dynamically sums bills + expenses + items by category so the budget page
    always reflects real spending rather than a cached value.
    """
    if home_id not in DB["homes"]:
        return jsonify({"error": "Home not found"}), 404

    def _compute_spent(category: str) -> float:
        cat = (category or "").strip().lower()
        total = 0.0
        for bill in DB["bills"].values():
            if bill.get("home_id") == home_id:
                if (bill.get("category") or "").strip().lower() == cat:
                    total += float(bill.get("total", 0) or 0)
        for expense in DB["expenses"].values():
            if expense.get("home_id") == home_id:
                if (expense.get("category") or "").strip().lower() == cat:
                    total += float(expense.get("amount", 0) or 0)
        for item in DB["items"].values():
            if item.get("home_id") == home_id:
                if (item.get("category") or "").strip().lower() == cat:
                    qty   = float(item.get("quantity", 0) or 0)
                    price = float(item.get("unit_price", 0) or 0)
                    total += qty * price
        return round(total, 2)

    result = []
    for budget in DB["budgets"].values():
        if budget.get("home_id") != home_id:
            continue
        b = dict(budget)
        b["current_balance"] = _compute_spent(b.get("category") or "")
        result.append(b)

    return jsonify({"budgets": result}), 200


# PATCH /api/budgets/<budget_id>/add-balance
@audit_bp.route("/budgets/<budget_id>/add-balance", methods=["PATCH"])
def add_budget_balance(budget_id):
    """FR-06: Add (or subtract) an amount from the current budget spending."""
    budget = DB["budgets"].get(budget_id)
    if not budget:
        return jsonify({"error": "Budget not found"}), 404

    data   = request.get_json(silent=True) or {}
    amount = data.get("amount")

    try:
        amount_val = float(amount)
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid amount"}), 400

    budget["current_balance"] = budget.get("current_balance", 0.0) + amount_val
    return jsonify({"budget": budget}), 200